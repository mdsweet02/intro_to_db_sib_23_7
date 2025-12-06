import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import pyodbc
import sys
from datetime import datetime
from tkinter import filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class DiplomaDBApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Дипломное проектирование - CRUD приложение с отчётами")
        self.root.geometry("1400x800")

        self.connection = None
        self.connect_to_db()

        self.create_widgets()
        self.create_menu()
        
    def connect_to_db(self):
        try:
            connection_string = (
                'DRIVER={SQL Server};'
                'SERVER=DESKTOP-DTSL47I\SQLEXPRESS;'
                'DATABASE=DiplomaDesign;'
                'Trusted_Connection=no;'
            )
            self.connection = pyodbc.connect(connection_string)
            print("Успешное подключение к базе данных")
        except pyodbc.Error as e:
            messagebox.showerror("Ошибка подклюения", f"Не удалось подключиться к базе данных:\n{str(e)}")
            sys.exit(1)
    
    def create_widgets(self):

        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill=tk.X)
        
        tables = [
            "Студенты", "Преподаватели", "Группы", "Специальности",
            "Проекты", "Результаты ГАК", "Расписание", "Банки", "ГАК Состав"
        ]
        
        for i, table in enumerate(tables):
            btn = ttk.Button(top_frame, text=table, 
                           command=lambda t=table: self.show_table(t))
            btn.grid(row=0, column=i, padx=5, pady=5)

        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(main_frame)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        crud_frame = ttk.Frame(self.root)
        crud_frame.pack(fill=tk.X, padx=10, pady=10)

        self.entry_fields = {}
        self.entry_frame = ttk.Frame(crud_frame)
        self.entry_frame.pack(fill=tk.X, pady=5)

        btn_frame = ttk.Frame(crud_frame)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="Добавить", command=self.add_record).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Обновить", command=self.update_record).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Удалить", command=self.delete_record).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Очистить", command=self.clear_form).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Обновить таблицу", command=self.refresh_table).pack(side=tk.LEFT, padx=5)

        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)

        self.current_table = "Students"
        self.show_table("Студенты")
    
    def create_menu(self):
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Выполнить SQL запрос", command=self.run_query)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.root.quit)

        reports_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Отчёты", menu=reports_menu)
        reports_menu.add_command(label="Отчёт 1: Результаты дипломирования", 
                               command=self.generate_report1)
        reports_menu.add_command(label="Отчёт 2: Оплата рецензентов", 
                               command=self.generate_report2)
        reports_menu.add_command(label="Отчёт 3: Состав ГАК", 
                               command=self.generate_report3)

        help_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Справка", menu=help_menu)
        help_menu.add_command(label="О программе", command=self.show_about)
    
    def show_table(self, table_name):
        """Отображение выбранной таблицы"""
        table_map = {
            "Студенты": "Students",
            "Преподаватели": "Professors",
            "Группы": "Groups",
            "Специальности": "Specialties",
            "Проекты": "Projects",
            "Результаты ГАК": "GAK_Results",
            "Расписание": "Projects_Defense_Schedule",
            "Банки": "Banks",
            "ГАК Состав": "GAK_Composition"
        }
        
        self.current_table = table_map.get(table_name)
        if not self.current_table:
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            cursor = self.connection.cursor()

            cursor.execute(f"SELECT * FROM {self.current_table} WHERE 1=0")
            columns = [column[0] for column in cursor.description]

            self.tree['columns'] = columns
            self.tree.heading('#0', text='ID')
            
            for col in columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100)

            cursor.execute(f"SELECT * FROM {self.current_table}")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                values = [str(item) if item is not None else "" for item in row]
                self.tree.insert('', 'end', text=str(i+1), values=values)

            self.create_form_fields(columns)
            
        except pyodbc.Error as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке таблицы:\n{str(e)}")
    
    def create_form_fields(self, columns):
        for widget in self.entry_frame.winfo_children():
            widget.destroy()
        
        self.entry_fields.clear()

        for i, col in enumerate(columns):
            label = ttk.Label(self.entry_frame, text=f"{col}:")
            label.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            
            entry = ttk.Entry(self.entry_frame, width=30)
            entry.grid(row=i, column=1, padx=5, pady=2)
            
            self.entry_fields[col] = entry
    
    def get_form_data(self):
        data = {}
        for col, entry in self.entry_fields.items():
            value = entry.get().strip()
            data[col] = value if value else None
        return data
    
    def add_record(self):
        try:
            data = self.get_form_data()

            columns = ', '.join(data.keys())
            placeholders = ', '.join(['?' for _ in data])
            values = list(data.values())
            
            cursor = self.connection.cursor()
            cursor.execute(f"INSERT INTO {self.current_table} ({columns}) VALUES ({placeholders})", values)
            self.connection.commit()
            
            messagebox.showinfo("Успех", "Запись успешно добавлена!")
            self.refresh_table()
            self.clear_form()
            
        except pyodbc.Error as e:
            messagebox.showerror("Ошибка", f"Ошибка при добавлении записи:\n{str(e)}")
    
    def update_record(self):
        try:
            selected_item = self.tree.selection()
            if not selected_item:
                messagebox.showwarning("Предупреждение", "Выберите запись для обновления!")
                return
            
            data = self.get_form_data()

            item_values = self.tree.item(selected_item[0])['values']
            cursor = self.connection.cursor()
            cursor.execute(f"SELECT * FROM {self.current_table} WHERE 1=0")
            columns = [column[0] for column in cursor.description]

            primary_key = columns[0]
            primary_key_value = item_values[0]

            set_clause = ', '.join([f"{col} = ?" for col in data.keys()])
            sql = f"UPDATE {self.current_table} SET {set_clause} WHERE {primary_key} = ?"
            
            values = list(data.values()) + [primary_key_value]
            
            cursor.execute(sql, values)
            self.connection.commit()
            
            messagebox.showinfo("Успех", "Запись успешно обновлена!")
            self.refresh_table()
            
        except pyodbc.Error as e:
            messagebox.showerror("Ошибка", f"Ошибка при обновлении записи:\n{str(e)}")
    
    def delete_record(self):
        try:
            selected_item = self.tree.selection()
            if not selected_item:
                messagebox.showwarning("Предупреждение", "Выберите запись для удаления!")
                return

            if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите удалить эту запись?"):
                return

            item_values = self.tree.item(selected_item[0])['values']
            cursor = self.connection.cursor()
            cursor.execute(f"SELECT * FROM {self.current_table} WHERE 1=0")
            columns = [column[0] for column in cursor.description]

            primary_key = columns[0]
            primary_key_value = item_values[0]

            cursor.execute(f"DELETE FROM {self.current_table} WHERE {primary_key} = ?", primary_key_value)
            self.connection.commit()
            
            messagebox.showinfo("Успех", "Запись успешно удалена!")
            self.refresh_table()
            self.clear_form()
            
        except pyodbc.Error as e:
            messagebox.showerror("Ошибка", f"Ошибка при удалении записи:\n{str(e)}")
    
    def clear_form(self):
        for entry in self.entry_fields.values():
            entry.delete(0, tk.END)
    
    def refresh_table(self):
        table_names = {
            "Students": "Студенты",
            "Professors": "Преподаватели",
            "Groups": "Группы",
            "Specialties": "Специальности",
            "Projects": "Проекты",
            "GAK_Results": "Результаты ГАК",
            "Projects_Defense_Schedule": "Расписание",
            "Banks": "Банки",
            "GAK_Composition": "ГАК Состав"
        }
        
        russian_name = table_names.get(self.current_table, "Студенты")
        self.show_table(russian_name)
    
    def on_tree_select(self, event):
        selected_item = self.tree.selection()
        if not selected_item:
            return
        
        item_values = self.tree.item(selected_item[0])['values']

        cursor = self.connection.cursor()
        cursor.execute(f"SELECT * FROM {self.current_table} WHERE 1=0")
        columns = [column[0] for column in cursor.description]
        
        self.clear_form()
        
        for col, value, entry in zip(columns, item_values, self.entry_fields.values()):
            if value is not None:
                entry.insert(0, str(value))
    
    def run_query(self):
        query_window = tk.Toplevel(self.root)
        query_window.title("Выполнить SQL запрос")
        query_window.geometry("600x400")
        
        ttk.Label(query_window, text="Введите SQL запрос:").pack(padx=10, pady=5)
        
        query_text = tk.Text(query_window, height=10)
        query_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        result_text = tk.Text(query_window, height=10, state=tk.DISABLED)
        result_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        def execute_query():
            query = query_text.get("1.0", tk.END).strip()
            if not query:
                return
            
            try:
                cursor = self.connection.cursor()
                cursor.execute(query)
                
                if query.strip().upper().startswith("SELECT"):
                    rows = cursor.fetchall()
                    result_text.config(state=tk.NORMAL)
                    result_text.delete("1.0", tk.END)
                    
                    for row in rows:
                        result_text.insert(tk.END, str(row) + "\n")
                    result_text.config(state=tk.DISABLED)
                else:
                    self.connection.commit()
                    messagebox.showinfo("Успех", "Запрос выполнен успешно!")
                    self.refresh_table()
                    
            except pyodbc.Error as e:
                messagebox.showerror("Ошибка", f"Ошибка выполнения запроса:\n{str(e)}")
        
        ttk.Button(query_window, text="Выполнить", command=execute_query).pack(pady=10)
    
    def show_about(self):
        messagebox.showinfo("О программе",
            "CRUD приложение для базы данных Дипломного проектирования\n\n"
            "Функции:\n"
            "1. Просмотр и редактирование всех таблиц\n"
            "2. Генерация отчётов в Excel\n"
            "3. Выполнение произвольных SQL запросов\n\n"
            "Требования:\n"
            "1. SQL Server\n"
            "2. Установленный pyodbc\n"
            "3. Созданная база данных DiplomaDesign")
    
    
    def generate_report1(self):
        date_dialog = tk.Toplevel(self.root)
        date_dialog.title("Выбор даты для отчёта 1")
        date_dialog.geometry("300x150")
        
        ttk.Label(date_dialog, text="Введите дату (ГГГГ-ММ-ДД):").pack(pady=10)
        
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_entry = ttk.Entry(date_dialog, textvariable=date_var, width=15)
        date_entry.pack(pady=5)
        
        def generate():
            report_date = date_var.get()
            
            try:
                datetime.strptime(report_date, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ГГГГ-ММ-ДД")
                return
            
            try:
                query = """
                SELECT 
                    s.student_surname + ' ' + s.student_first_name + ' ' + 
                    ISNULL(s.student_patronymic, '') AS ФИО_студента,
                    g.group_name AS Группа,
                    p.professor_surname + ' ' + p.professor_first_name + ' ' + 
                    ISNULL(p.professor_patronymic, '') AS ФИО_руководителя,
                    r.reviewer_surname + ' ' + r.reviewer_first_name + ' ' + 
                    ISNULL(r.reviewer_patronymic, '') AS ФИО_рецензента,
                    gr.grade AS Оценка,
                    gr.final_assessment_type AS Тип_аттестации
                FROM Students s
                JOIN Groups g ON s.group_code = g.group_code
                JOIN Projects pr ON s.student_code = pr.student_code
                JOIN Professors p ON pr.professor_code = p.professor_code
                JOIN GAK_Reviewers r ON pr.reviewer_code = r.iin
                JOIN GAK_Results gr ON s.student_code = gr.student_code
                WHERE gr.defense_date = ?
                ORDER BY s.student_surname, s.student_first_name
                """
                
                cursor = self.connection.cursor()
                cursor.execute(query, report_date)
                rows = cursor.fetchall()
                
                if not rows:
                    messagebox.showinfo("Нет данных", f"Нет данных за дату {report_date}")
                    date_dialog.destroy()
                    return
                
                columns = ['ФИО студента', 'Группа', 'ФИО руководителя', 
                          'ФИО рецензента', 'Оценка', 'Тип аттестации']
                df = pd.DataFrame.from_records(rows, columns=columns)

                self.export_report_to_excel(df, 
                    f"Отчет_по_результатам_дипломирования_{report_date}",
                    f"Отчёт по результатам дипломирования на {report_date}")
                
                date_dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при генерации отчёта:\n{str(e)}")
        
        ttk.Button(date_dialog, text="Сгенерировать отчёт", 
                  command=generate).pack(pady=10)
    
    def generate_report2(self):
        try:
            query = """
            SELECT 
                r.reviewer_surname + ' ' + r.reviewer_first_name + ' ' + 
                ISNULL(r.reviewer_patronymic, '') AS ФИО_рецензента,
                r.iin AS ИИН,
                r.work_place AS Место_работы,
                r.job_title AS Должность,
                COUNT(pr.student_code) AS Количество_рецензий,
                COUNT(pr.student_code) * 2000 AS Сумма_оплаты_тенге
            FROM GAK_Reviewers r
            LEFT JOIN Projects pr ON r.iin = pr.reviewer_code
            GROUP BY r.iin, r.reviewer_surname, r.reviewer_first_name, 
                     r.reviewer_patronymic, r.work_place, r.job_title
            ORDER BY Количество_рецензий DESC, r.reviewer_surname
            """
            
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()

            columns = ['ФИО рецензента', 'ИИН', 'Место работы', 
                      'Должность', 'Количество рецензий', 'Сумма оплаты (тенге)']
            df = pd.DataFrame.from_records(rows, columns=columns)

            total_reviews = df['Количество рецензий'].sum()
            total_payment = df['Сумма оплаты (тенге)'].sum()
            
            totals_df = pd.DataFrame([{
                'ФИО рецензента': 'ИТОГО:',
                'ИИН': '',
                'Место работы': '',
                'Должность': '',
                'Количество рецензий': total_reviews,
                'Сумма оплаты (тенге)': total_payment
            }])
            
            df = pd.concat([df, totals_df], ignore_index=True)

            self.export_report_to_excel(df, 
                "Список_оплаты_труда_рецензентов",
                "Список оплаты труда рецензентов\nСтоимость одной рецензии: 2000 тенге")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при генерации отчёта:\n{str(e)}")
    
    def generate_report3(self):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT specialty_code, specialty_name FROM Specialties ORDER BY specialty_name")
            specialties = cursor.fetchall()
            
            if not specialties:
                messagebox.showwarning("Нет данных", "Нет доступных специальностей")
                return

            choice_dialog = tk.Toplevel(self.root)
            choice_dialog.title("Выбор специальности для отчёта 3")
            choice_dialog.geometry("400x150")
            
            ttk.Label(choice_dialog, text="Выберите специальность:").pack(pady=10)
            
            specialty_var = tk.StringVar()
            specialty_names = [f"{code} - {name}" for code, name in specialties]
            specialty_var.set(specialty_names[0])
            
            specialty_combo = ttk.Combobox(choice_dialog, 
                                          textvariable=specialty_var,
                                          values=specialty_names,
                                          state="readonly",
                                          width=40)
            specialty_combo.pack(pady=10)
            
            def generate():
                selected = specialty_var.get()
                specialty_code = int(selected.split(" - ")[0])
                specialty_name = selected.split(" - ")[1]
                
                try:
                    query = """
                    SELECT 
                        s.specialty_name AS Специальность,
                        gc.iin AS ИИН_члена_ГАК,
                        r.reviewer_surname + ' ' + r.reviewer_first_name + ' ' + 
                        ISNULL(r.reviewer_patronymic, '') AS ФИО_члена_ГАК,
                        r.work_place AS Место_работы,
                        r.job_title AS Должность,
                        gc.functions AS Функция_в_ГАК,
                        r.graduates_number AS Количество_выпускников
                    FROM GAK_Composition gc
                    JOIN Specialties s ON gc.specialty_code = s.specialty_code
                    JOIN GAK_Reviewers r ON gc.iin = r.iin
                    WHERE gc.specialty_code = ?
                    ORDER BY 
                        CASE gc.functions 
                            WHEN 'председатель' THEN 1
                            WHEN 'секретарь' THEN 2
                            WHEN 'член ГАКа' THEN 3
                            ELSE 4
                        END,
                        r.reviewer_surname
                    """
                    
                    cursor = self.connection.cursor()
                    cursor.execute(query, specialty_code)
                    rows = cursor.fetchall()
                    
                    if not rows:
                        messagebox.showinfo("Нет данных", f"Для специальности '{specialty_name}' не найден состав ГАК")
                        choice_dialog.destroy()
                        return

                    columns = ['Специальность', 'ИИН члена ГАК', 'ФИО члена ГАК', 
                              'Место работы', 'Должность', 'Функция в ГАК', 'Количество выпускников']
                    df = pd.DataFrame.from_records(rows, columns=columns)

                    self.export_report_to_excel(df, 
                        f"Состав_ГАК_{specialty_name.replace(' ', '_')}",
                        f"Состав ГАК по специальности: {specialty_name}")
                    
                    choice_dialog.destroy()
                    
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка при генерации отчёта:\n{str(e)}")
            
            ttk.Button(choice_dialog, text="Сгенерировать отчёт", 
                      command=generate).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при получении списка специальностей:\n{str(e)}")
    
    def export_report_to_excel(self, df, default_filename, title):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{default_filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Отчёт')

                workbook = writer.book
                worksheet = writer.sheets['Отчёт']

                self.format_excel_worksheet(worksheet, df, title)
                
            messagebox.showinfo("Успех", f"Отчёт успешно экспортирован в:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать в Excel:\n{str(e)}")
    
    def format_excel_worksheet(self, worksheet, df, title):
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        worksheet.insert_rows(1, 2)
        worksheet.merge_cells('A1:G1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

            column_letter = get_column_letter(col)
            max_length = max(
                df.iloc[:, col-1].astype(str).apply(len).max(),
                len(df.columns[col-1])
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        for row in range(4, len(df) + 4):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border

                if 'Оценка' in df.columns[col-1] or 'Количество' in df.columns[col-1] or 'Сумма' in df.columns[col-1]:
                    cell.alignment = Alignment(horizontal='right')

                if 'ИТОГО:' in str(cell.value):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        worksheet.auto_filter.ref = f"A3:{get_column_letter(len(df.columns))}{len(df) + 3}"


def main():
    root = tk.Tk()
    app = DiplomaDBApp(root)
    root.mainloop()

if __name__ == "__main__":
    try:
        import pandas
        import openpyxl
    except ImportError as e:
        print(f"Ошибка: Не установлены необходимые библиотеки: {e}")
        print("Установите их с помощью: pip install pandas openpyxl")
        sys.exit(1)
    
    main()